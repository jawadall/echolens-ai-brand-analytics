"""
Report Generators for Echo Lens
Generates CSV, PDF, and Excel exports with comprehensive analytics
"""
import io
import csv
import re
import logging
from typing import Dict, List
from django.utils import timezone
from collections import Counter

logger = logging.getLogger(__name__)


def _md_to_rl(text: str) -> str:
    """Convert Gemini markdown to ReportLab XML markup."""
    if not text:
        return ''
    # Escape XML entities first
    text = text.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
    # **bold** → <b>bold</b>
    text = re.sub(r'\*\*(.+?)\*\*', r'<b>\1</b>', text)
    # [URGENT] etc → colored bold tags
    tag_colors = {'URGENT': '#ef4444', 'HIGH': '#f97316', 'MEDIUM': '#eab308', 'LOW': '#22c55e', 'LONG-TERM': '#8b5cf6'}
    for tag, color in tag_colors.items():
        text = text.replace(f'[{tag}]', f'<font color="{color}"><b>[{tag}]</b></font> ')
    # 'quoted' → italic
    text = re.sub(r"'([^']{2,80})'", r'<i>&#8216;\1&#8217;</i>', text)
    return text


class CSVGenerator:
    """Generate CSV exports"""

    @staticmethod
    def generate_posts_csv(posts: List[Dict]) -> str:
        """Generate CSV from posts data"""
        output = io.StringIO()
        writer = csv.writer(output)

        headers = [
            'ID', 'Platform', 'Author', 'Username', 'Content',
            'Sentiment', 'Score', 'Likes', 'Shares', 'Comments',
            'Views', 'Topics', 'Posted At', 'URL'
        ]
        writer.writerow(headers)

        for post in posts:
            writer.writerow([
                post.get('id', ''),
                post.get('platform', ''),
                post.get('author_name', ''),
                post.get('author_username', ''),
                post.get('content', '')[:500],
                post.get('sentiment', ''),
                post.get('sentiment_score', ''),
                post.get('likes', 0),
                post.get('shares', 0),
                post.get('comments', 0),
                post.get('views', 0),
                ', '.join(post.get('topics', [])),
                str(post.get('posted_at', '')),
                post.get('url', ''),
            ])

        return output.getvalue()

    @staticmethod
    def generate_analytics_csv(analytics_data: Dict) -> str:
        """Generate CSV from analytics data"""
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(['Echo Lens Analytics Report'])
        writer.writerow(['Generated:', timezone.now().isoformat()])
        writer.writerow([])

        # Summary
        writer.writerow(['Summary'])
        writer.writerow(['Metric', 'Value'])
        writer.writerow(['Total Posts', analytics_data.get('total_posts', 0)])
        writer.writerow(['Positive', analytics_data.get('positive', 0)])
        writer.writerow(['Neutral', analytics_data.get('neutral', 0)])
        writer.writerow(['Negative', analytics_data.get('negative', 0)])
        writer.writerow(['Average Sentiment', analytics_data.get('avg_sentiment', 0)])
        writer.writerow([])

        # Daily data if available
        if 'daily_data' in analytics_data:
            writer.writerow(['Daily Breakdown'])
            writer.writerow(['Date', 'Total', 'Positive', 'Neutral', 'Negative', 'Avg Sentiment'])
            for day in analytics_data['daily_data']:
                writer.writerow([
                    day.get('date', ''),
                    day.get('total', 0),
                    day.get('positive', 0),
                    day.get('neutral', 0),
                    day.get('negative', 0),
                    day.get('avg_sentiment', 0),
                ])

        return output.getvalue()


class PDFGenerator:
    """Generate comprehensive multi-page PDF reports"""

    @staticmethod
    def generate_summary_pdf(brand_name: str, analytics_data: Dict, summary_data: Dict = None) -> bytes:
        """Generate professional multi-page PDF summary report"""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch, cm
            from reportlab.platypus import (
                SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
                PageBreak, HRFlowable
            )
            from reportlab.graphics.shapes import Drawing
            from reportlab.graphics.charts.piecharts import Pie
            from reportlab.graphics.charts.barcharts import VerticalBarChart
            from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

            buffer = io.BytesIO()
            doc = SimpleDocTemplate(
                buffer,
                pagesize=A4,
                topMargin=1 * cm,
                bottomMargin=1.5 * cm,
                leftMargin=2 * cm,
                rightMargin=2 * cm
            )
            styles = getSampleStyleSheet()
            story = []

            # ═══════════════ CUSTOM STYLES ═══════════════
            brand_color = colors.HexColor('#6366f1')  # Primary brand purple
            dark_bg = colors.HexColor('#1e1b4b')
            accent = colors.HexColor('#818cf8')
            success = colors.HexColor('#10b981')
            warning = colors.HexColor('#f59e0b')
            danger = colors.HexColor('#f43f5e')

            cover_title_style = ParagraphStyle(
                'CoverTitle', parent=styles['Title'],
                fontSize=36, textColor=brand_color, spaceAfter=10,
                alignment=TA_CENTER, fontName='Helvetica-Bold'
            )
            cover_subtitle_style = ParagraphStyle(
                'CoverSubtitle', parent=styles['Normal'],
                fontSize=18, textColor=colors.HexColor('#64748b'),
                alignment=TA_CENTER, spaceAfter=5
            )
            section_heading = ParagraphStyle(
                'SectionHeading', parent=styles['Heading1'],
                fontSize=20, textColor=brand_color, spaceBefore=20,
                spaceAfter=12, fontName='Helvetica-Bold',
                borderWidth=2, borderColor=brand_color, borderPadding=5
            )
            sub_heading = ParagraphStyle(
                'SubHeading', parent=styles['Heading2'],
                fontSize=14, textColor=colors.HexColor('#334155'),
                spaceBefore=12, spaceAfter=8, fontName='Helvetica-Bold'
            )
            body_style = ParagraphStyle(
                'BodyText2', parent=styles['Normal'],
                fontSize=10, leading=14, spaceBefore=4, spaceAfter=4,
                textColor=colors.HexColor('#1e293b')
            )
            small_style = ParagraphStyle(
                'SmallText', parent=styles['Normal'],
                fontSize=8, textColor=colors.HexColor('#94a3b8')
            )

            # ═══════════════ PAGE 1: COVER PAGE ═══════════════
            story.append(Spacer(1, 2 * inch))
            story.append(Paragraph("ECHO LENS", cover_title_style))
            story.append(Paragraph("Brand Monitoring & Sentiment Analysis Report", cover_subtitle_style))
            story.append(Spacer(1, 0.5 * inch))
            story.append(HRFlowable(
                width="60%", thickness=3, color=brand_color,
                spaceAfter=20, spaceBefore=10
            ))
            story.append(Spacer(1, 0.3 * inch))
            story.append(Paragraph(f'<font size="24" color="#1e293b"><b>{brand_name}</b></font>', 
                                    ParagraphStyle('BrandName', alignment=TA_CENTER)))
            story.append(Spacer(1, 0.3 * inch))

            now = timezone.now()
            cover_info = [
                ['Report Date', now.strftime('%B %d, %Y')],
                ['Generated At', now.strftime('%I:%M %p')],
                ['Total Posts Analyzed', f"{analytics_data.get('total_posts', 0):,}"],
                ['Analysis Period', f"Last {analytics_data.get('days', 30)} days"],
            ]
            cover_table = Table(cover_info, colWidths=[2.5 * inch, 2.5 * inch])
            cover_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 11),
                ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#64748b')),
                ('TEXTCOLOR', (1, 0), (1, -1), colors.HexColor('#1e293b')),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
            ]))
            story.append(cover_table)

            story.append(Spacer(1, 1.5 * inch))
            story.append(Paragraph(
                '<i>Generated by Echo Lens — AI-Powered Brand Monitoring Platform</i>',
                ParagraphStyle('Footer', parent=small_style, alignment=TA_CENTER, fontSize=9)
            ))
            story.append(PageBreak())

            # ═══════════════ PAGE 2: EXECUTIVE SUMMARY ═══════════════
            story.append(Paragraph("1. Executive Summary", section_heading))

            total = analytics_data.get('total_posts', 0)
            positive = analytics_data.get('positive', 0)
            neutral = analytics_data.get('neutral', 0)
            negative = analytics_data.get('negative', 0)
            avg_sent = analytics_data.get('avg_sentiment', 0)

            pos_pct = (positive / total * 100) if total > 0 else 0
            neg_pct = (negative / total * 100) if total > 0 else 0
            neu_pct = (neutral / total * 100) if total > 0 else 0

            # Key Metrics Cards
            metrics_data = [
                ['Total Posts', 'Positive', 'Neutral', 'Negative', 'Avg Score'],
                [f'{total:,}', f'{positive:,}\n({pos_pct:.1f}%)',
                 f'{neutral:,}\n({neu_pct:.1f}%)', f'{negative:,}\n({neg_pct:.1f}%)',
                 f'{avg_sent:.3f}']
            ]
            metrics_table = Table(metrics_data, colWidths=[1.2 * inch] * 5)
            metrics_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), brand_color),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('FONTSIZE', (0, 1), (-1, 1), 12),
                ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
                ('TOPPADDING', (0, 0), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e2e8f0')),
                ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#f8fafc')),
            ]))
            story.append(metrics_table)
            story.append(Spacer(1, 20))

            # AI Summary
            if summary_data and summary_data.get('summary'):
                story.append(Paragraph("AI-Generated Analysis", sub_heading))
                story.append(Paragraph(_md_to_rl(summary_data['summary']), body_style))
                story.append(Spacer(1, 10))

            # Sentiment Distribution Pie Chart
            story.append(Paragraph("Sentiment Distribution", sub_heading))
            if total > 0:
                d = Drawing(300, 200)
                pie = Pie()
                pie.x = 80
                pie.y = 15
                pie.width = 140
                pie.height = 140
                pie.data = [positive, neutral, negative]
                pie.labels = [
                    f'Positive ({pos_pct:.1f}%)',
                    f'Neutral ({neu_pct:.1f}%)',
                    f'Negative ({neg_pct:.1f}%)'
                ]
                pie.slices[0].fillColor = success
                pie.slices[1].fillColor = warning
                pie.slices[2].fillColor = danger
                pie.slices.strokeColor = colors.white
                pie.slices.strokeWidth = 2
                d.add(pie)
                story.append(d)
            story.append(Spacer(1, 15))

            story.append(PageBreak())

            # ═══════════════ PAGE 3: INSIGHTS & ANALYSIS ═══════════════
            story.append(Paragraph("2. Key Insights &amp; Analysis", section_heading))

            if summary_data and summary_data.get('key_insights'):
                for i, insight in enumerate(summary_data['key_insights'], 1):
                    story.append(Paragraph(
                        f'<font color="#6366f1"><b>{i}.</b></font> {_md_to_rl(insight)}',
                        body_style
                    ))
                    story.append(Spacer(1, 4))
                story.append(Spacer(1, 10))

            if summary_data and summary_data.get('what_users_like'):
                story.append(Paragraph(
                    '<font color="#10b981">&#9650;</font> Positive Drivers', sub_heading
                ))
                story.append(Paragraph(_md_to_rl(summary_data['what_users_like']), body_style))
                story.append(Spacer(1, 10))

            if summary_data and summary_data.get('what_users_dislike'):
                story.append(Paragraph(
                    '<font color="#f43f5e">&#9660;</font> Negative Drivers', sub_heading
                ))
                story.append(Paragraph(_md_to_rl(summary_data['what_users_dislike']), body_style))
                story.append(Spacer(1, 10))

            if summary_data and summary_data.get('platform_analysis'):
                story.append(Paragraph(
                    '<font color="#6366f1">&#9673;</font> Platform-Specific Observations', sub_heading
                ))
                story.append(Paragraph(_md_to_rl(summary_data['platform_analysis']), body_style))
                story.append(Spacer(1, 10))

            # ═══════════════ PAGE 4: PLATFORM BREAKDOWN ═══════════════
            if analytics_data.get('platform_data'):
                story.append(PageBreak())
                story.append(Paragraph("3. Platform Breakdown", section_heading))

                platform_headers = ['Platform', 'Posts', 'Positive %', 'Negative %', 'Avg Score']
                platform_rows = [platform_headers]
                for p_data in analytics_data['platform_data']:
                    p_total = p_data.get('total', 0)
                    p_pos = p_data.get('positive', 0)
                    p_neg = p_data.get('negative', 0)
                    platform_rows.append([
                        p_data.get('platform', '').title(),
                        str(p_total),
                        f"{(p_pos/p_total*100) if p_total > 0 else 0:.1f}%",
                        f"{(p_neg/p_total*100) if p_total > 0 else 0:.1f}%",
                        f"{p_data.get('avg_sentiment', 0):.3f}",
                    ])

                if len(platform_rows) > 1:
                    pt = Table(platform_rows, colWidths=[1.5*inch, 1*inch, 1.2*inch, 1.2*inch, 1.2*inch])
                    pt.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#334155')),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e2e8f0')),
                        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#f8fafc'), colors.white]),
                        ('PADDING', (0, 0), (-1, -1), 8),
                    ]))
                    story.append(pt)
                    story.append(Spacer(1, 20))

            # ═══════════════ PAGE 5: TOPICS ═══════════════
            if analytics_data.get('top_topics'):
                story.append(Paragraph("4. Trending Topics", section_heading))
                topics_data = [['Rank', 'Topic', 'Mentions']]
                for i, topic in enumerate(analytics_data['top_topics'][:15], 1):
                    topics_data.append([
                        str(i),
                        topic.get('topic', ''),
                        str(topic.get('count', 0))
                    ])

                topics_table = Table(topics_data, colWidths=[0.8*inch, 3.5*inch, 1.2*inch])
                topics_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4338ca')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('ALIGN', (0, 0), (0, -1), 'CENTER'),
                    ('ALIGN', (2, 0), (2, -1), 'CENTER'),
                    ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e2e8f0')),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#f8fafc'), colors.white]),
                    ('PADDING', (0, 0), (-1, -1), 6),
                ]))
                story.append(topics_table)
                story.append(Spacer(1, 20))

            # ═══════════════ PAGE 6: RECOMMENDATIONS ═══════════════
            if summary_data and summary_data.get('recommendations'):
                story.append(PageBreak())
                story.append(Paragraph("5. Strategic Recommendations", section_heading))
                for i, rec in enumerate(summary_data['recommendations'], 1):
                    story.append(Paragraph(
                        f'<font color="#6366f1"><b>#{i}</b></font> {_md_to_rl(rec)}',
                        body_style
                    ))
                    story.append(Spacer(1, 8))

            # ═══════════════ FOOTER ON ALL PAGES ═══════════════
            story.append(Spacer(1, 40))
            story.append(HRFlowable(
                width="100%", thickness=1, color=colors.HexColor('#e2e8f0')
            ))
            story.append(Paragraph(
                f'<i>Echo Lens Report — {brand_name} — Generated {now.strftime("%Y-%m-%d %H:%M")} — '
                f'Confidential</i>',
                ParagraphStyle('FooterFinal', parent=small_style,
                               alignment=TA_CENTER, fontSize=8)
            ))

            doc.build(story)
            return buffer.getvalue()

        except ImportError as e:
            logger.error(f"ReportLab not installed: {e}")
            content = f"""
Echo Lens - Brand Report
========================
Brand: {brand_name}
Generated: {timezone.now().isoformat()}

Overview:
- Total Posts: {analytics_data.get('total_posts', 0)}
- Positive: {analytics_data.get('positive', 0)}
- Neutral: {analytics_data.get('neutral', 0)}
- Negative: {analytics_data.get('negative', 0)}
- Avg Sentiment: {analytics_data.get('avg_sentiment', 0):.3f}
"""
            return content.encode('utf-8')
        except Exception as e:
            logger.error(f"PDF generation error: {e}")
            raise


class ExcelGenerator:
    """Generate Excel exports"""

    @staticmethod
    def generate_posts_excel(posts: List[Dict]) -> bytes:
        """Generate Excel from posts data"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = Workbook()
            ws = wb.active
            ws.title = "Posts Data"

            # Headers
            headers = [
                'ID', 'Platform', 'Author', 'Username',
                'Content', 'Sentiment', 'Score',
                'Likes', 'Shares', 'Comments', 'Views',
                'Topics', 'Language', 'Posted At', 'URL'
            ]

            # Style headers
            header_font = Font(bold=True, color='FFFFFF', size=11)
            header_fill = PatternFill(start_color='6366F1', end_color='6366F1', fill_type='solid')
            thin_border = Border(
                left=Side(style='thin', color='E2E8F0'),
                right=Side(style='thin', color='E2E8F0'),
                top=Side(style='thin', color='E2E8F0'),
                bottom=Side(style='thin', color='E2E8F0')
            )

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border

            # Sentiment color fills
            pos_fill = PatternFill(start_color='DCFCE7', end_color='DCFCE7', fill_type='solid')
            neg_fill = PatternFill(start_color='FFE4E6', end_color='FFE4E6', fill_type='solid')
            neu_fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')

            # Data
            for row, post in enumerate(posts, 2):
                ws.cell(row=row, column=1, value=post.get('id', ''))
                ws.cell(row=row, column=2, value=post.get('platform', ''))
                ws.cell(row=row, column=3, value=post.get('author_name', ''))
                ws.cell(row=row, column=4, value=post.get('author_username', ''))
                ws.cell(row=row, column=5, value=post.get('content', '')[:500])
                sent_cell = ws.cell(row=row, column=6, value=post.get('sentiment', ''))
                ws.cell(row=row, column=7, value=post.get('sentiment_score', ''))
                ws.cell(row=row, column=8, value=post.get('likes', 0))
                ws.cell(row=row, column=9, value=post.get('shares', 0))
                ws.cell(row=row, column=10, value=post.get('comments', 0))
                ws.cell(row=row, column=11, value=post.get('views', 0))
                ws.cell(row=row, column=12, value=', '.join(post.get('topics', [])))
                ws.cell(row=row, column=13, value=post.get('language', 'en'))
                ws.cell(row=row, column=14, value=str(post.get('posted_at', '')))
                ws.cell(row=row, column=15, value=post.get('url', ''))

                # Color-code sentiment
                sentiment = post.get('sentiment', '')
                if sentiment == 'positive':
                    sent_cell.fill = pos_fill
                elif sentiment == 'negative':
                    sent_cell.fill = neg_fill
                elif sentiment == 'neutral':
                    sent_cell.fill = neu_fill

            # Adjust column widths
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 18
            ws.column_dimensions['D'].width = 18
            ws.column_dimensions['E'].width = 60
            ws.column_dimensions['F'].width = 12
            ws.column_dimensions['G'].width = 10
            ws.column_dimensions['L'].width = 30
            ws.column_dimensions['M'].width = 10
            ws.column_dimensions['N'].width = 22
            ws.column_dimensions['O'].width = 50

            # Freeze header row
            ws.freeze_panes = 'A2'

            buffer = io.BytesIO()
            wb.save(buffer)
            return buffer.getvalue()

        except ImportError:
            logger.error("openpyxl not installed")
            raise
